import os
import time
import pandas as pd
import sqlite3
from queue import Queue
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from rapidfuzz import fuzz, process
import threading
import queue
from openpyxl import load_workbook
from datetime import datetime
import re

WATCH_FOLDER = r'enter/path/to/be/watched/here'
DB_PATH = 'row_clean.db'
TABLE_NAME = 'person'

file_queue = Queue()
preprocessing_queue = Queue()
processed_files = set()
observer = None
log_queue = queue.Queue()
merge_ready_files = []
MERGED_OUTPUT_PATH = os.path.join(WATCH_FOLDER, "output.xlsx")
session_normalized_queue = Queue()

# Session tracking for merge logic
session_files = set()  # Track files processed in current session
session_lock = threading.Lock()

file_counter = 0
cleaned_counter = 0
counter_lock = threading.Lock()

def log(msg):
    print(msg)
    log_queue.put(msg)

def get_log_queue():
    return log_queue

def get_name_from_filename(filename):
    """Enhanced filename matching using substring and fuzzy matching"""
    base = os.path.splitext(os.path.basename(filename))[0]
    
    # Clean the filename: remove special characters, numbers, and normalize
    cleaned_base = re.sub(r'[^a-zA-Z\s]', ' ', base)
    cleaned_base = re.sub(r'\s+', ' ', cleaned_base).strip().lower()
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT name FROM person")
    names = [row[0] for row in cursor.fetchall()]
    conn.close()

    log(f"🔍 Matching filename: '{base}' (cleaned: '{cleaned_base}') against {len(names)} names from DB")
    
    # First try: exact substring matching
    for name in names:
        name_lower = name.lower()
        if name_lower in cleaned_base or any(word in name_lower for word in cleaned_base.split()):
            log(f"🎯 Substring match found: '{cleaned_base}' → '{name}'")
            return name
    
    # Second try: fuzzy matching with lower threshold since we're doing substring first
    names_lower = [name.lower() for name in names]
    match, score, _ = process.extractOne(cleaned_base, names_lower, scorer=fuzz.ratio)
    if score >= 70:  # Lowered threshold since substring didn't work
        original_name = next(name for name in names if name.lower() == match)
        log(f"🔍 Fuzzy match: '{cleaned_base}' → '{original_name}' (score: {score})")
        return original_name
    else:
        log(f"⚠️ No good match for '{cleaned_base}' (best score: {score})")
        return None

def load_mappings_from_db(name):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(f"SELECT col_name, norm_col_name FROM {TABLE_NAME} WHERE LOWER(name)=?", (name.lower(),))
    mappings = {col.lower(): norm for col, norm in cursor.fetchall()}
    conn.close()
    log(f"🗺️ Loaded mappings for supplier: {name} → {len(mappings)} columns")
    return mappings

def get_db_column_names(name):
    """Get all column names from database for header detection"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(f"SELECT col_name FROM {TABLE_NAME} WHERE LOWER(name)=?", (name.lower(),))
    col_names = [row[0].lower().strip() for row in cursor.fetchall()]
    conn.close()
    return col_names

def detect_header_row_from_db(file_path, name):
    """Enhanced header detection using database column names"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    db_columns = get_db_column_names(name)
    
    if not db_columns:
        log(f"❌ No column names found in DB for supplier: {name}")
        return pd.DataFrame()
    
    log(f"🔍 Looking for header row using DB columns: {db_columns}")
    
    dataframes = []
    for sheet in wb.worksheets:
        log(f"🧪 Scanning sheet: {sheet.title}")
        rows = list(sheet.iter_rows(values_only=True))
        
        header_row_idx = None
        header_row = None
        
        # Look for header row by matching with database column names
        for i, row in enumerate(rows):
            if row is None:
                continue
                
            # Clean and normalize row values
            row_values = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            non_empty_values = [val for val in row_values if val]
            
            if len(non_empty_values) < 3:  # Skip rows with too few values
                continue
            
            # Count matches with database columns
            matches = 0
            for val in non_empty_values:
                if val in db_columns:
                    matches += 1
            
            # If we find a good match ratio, this is likely the header
            match_ratio = matches / len(non_empty_values) if non_empty_values else 0
            if match_ratio >= 0.3:  # At least 30% of columns match DB
                header_row_idx = i
                header_row = [str(cell).strip() if cell is not None else "" for cell in row]
                log(f"✅ Found header row at index {i} with {matches}/{len(non_empty_values)} matches ({match_ratio:.1%})")
                break
        
        if header_row_idx is not None and header_row:
            # Extract data rows after header
            data_rows = []
            for j in range(header_row_idx + 1, len(rows)):
                if rows[j] is not None:
                    data_rows.append(rows[j])
            
            if data_rows:
                df = pd.DataFrame(data_rows, columns=header_row)
                log(f"📊 Created dataframe for sheet: {sheet.title} with shape {df.shape}")
                dataframes.append(df)
    
    if dataframes:
        combined_df = pd.concat(dataframes, ignore_index=True)
        log(f"📊 Combined dataframe shape: {combined_df.shape}")
        return combined_df
    else:
        log(f"❌ No valid header rows found in {file_path}")
        return pd.DataFrame()

def read_excel_safely(file_path, min_non_na=5):
    """Fallback method - kept for compatibility"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    dataframes = []

    for sheet in wb.worksheets:
        log(f"🧪 Scanning sheet: {sheet.title}")
        rows = sheet.iter_rows(values_only=True)
        header_row = None
        data_rows = []

        for i, row in enumerate(rows):
            if row is None:
                continue
            if header_row is None and sum(cell is not None for cell in row[:min_non_na]) >= min_non_na:
                header_row = [str(cell).strip() if cell is not None else "" for cell in row]
                log(f"✅ Detected header row: {header_row}")
            elif header_row:
                data_rows.append(row)

        if header_row and data_rows:
            df = pd.DataFrame(data_rows, columns=header_row)
            log(f"📊 Created dataframe for sheet: {sheet.title} with shape {df.shape}")
            dataframes.append(df)

    if dataframes:
        return pd.concat(dataframes, ignore_index=True)
    else:
        log(f"❌ No valid sheets found in {file_path}")
        return pd.DataFrame()

def split_measurement_columns(df, name):
    rules = {
        ('abc', 'Measurements'): { 
            'pattern': 'a-b*c',
            'new_cols': ['Min', 'Max', 'Height']
        },
        
    }

    df.columns = [c.strip() for c in df.columns]
    for (supplier, norm_col), rule in rules.items():
        match_col = next((col for col in df.columns if col.lower() == norm_col.lower()), None)
        if name.lower() == supplier.lower() and match_col:
            def parse_expr(val):
                try:
                    val = str(val)
                    if rule['pattern'] == 'a-b*c':
                        a, rest = val.split('-')
                        b, c = rest.split('*')
                        return float(a), float(b), float(c)
                    elif rule['pattern'] == 'a*b-c':
                        a, rest = val.split('*')
                        b, c = rest.split('-')
                        return float(a), float(b), float(c)
                except:
                    return None, None, None

            new_cols = rule['new_cols']
            df[new_cols] = df[match_col].apply(lambda x: pd.Series(parse_expr(x)))
            df.drop(columns=[match_col], inplace=True)

    return df

def apply_combined_filters(df, db_path='filter.db'):
    log(f"🧽 Starting cleaning for file with shape: {df.shape}")
    df.columns = [c.strip() for c in df.columns]

    FIXED_WHITELIST = {
        'Color': {'red', 'blue', 'Green'},
        
    }

    for col, allowed_values in FIXED_WHITELIST.items():
        match_col = next((c for c in df.columns if c.lower() == col.lower()), None)
        if match_col:
            df = df[df[match_col].astype(str).str.strip().str.upper().isin({v.upper() for v in allowed_values})]
    log(f"🧼 After FIXED_WHITELIST, shape is: {df.shape}")

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT col_name, raw_value FROM data_filter")
    rows = cursor.fetchall()
    conn.close()

    filter_dict = {}
    for col, raw in rows:
        filter_dict.setdefault(col.lower(), set()).add(raw.lower())

    log(f"🔍 DB-based filters loaded for columns: {list(filter_dict.keys())}")
    for col in filter_dict:
        match_col = next((c for c in df.columns if c.lower() == col), None)
        if match_col:
            df = df[df[match_col].astype(str).str.lower().isin(filter_dict[col])]
    log(f"🧼 After DB filters, final shape: {df.shape}")

    return df

def merge_session_files():
    """Merge only files from current session using existing queues"""
    log("🔗 Attempting to merge files from current session...")
    
    with session_lock:
        if not session_files:
            log("⚠️ No files in current session to merge.")
            return
        
        session_files_copy = session_files.copy()
    
    dataframes = []
    successfully_merged = []
    
    for file_path in session_files_copy:
        # Convert to normalized filename
        normalized_file = os.path.splitext(file_path)[0] + '_normalized.xlsx'
        
        if os.path.exists(normalized_file):
            try:
                df = pd.read_excel(normalized_file)
                if not df.empty:
                    dataframes.append(df)
                    successfully_merged.append(os.path.basename(file_path))
                    log(f"✅ Added to merge: {os.path.basename(normalized_file)} (shape: {df.shape})")
                else:
                    log(f"⚠️ Empty dataframe in: {normalized_file}")
            except Exception as e:
                log(f"❌ Failed to read {normalized_file} for merge: {e}")
        else:
            log(f"⚠️ Normalized file not found: {normalized_file}")

    if dataframes:
        try:
            final_df = pd.concat(dataframes, ignore_index=True)
            final_df.to_excel(MERGED_OUTPUT_PATH, index=False)
            log(f"✅ Session merge completed! Final file: {MERGED_OUTPUT_PATH}")
            log(f"📊 Merged {len(successfully_merged)} files with total shape: {final_df.shape}")
            log(f"📁 Files merged: {', '.join(successfully_merged)}")
            
            # Clear session files after successful merge
            with session_lock:
                session_files.clear()
                
        except Exception as e:
            log(f"❌ Error during merge: {e}")
    else:
        log("⚠️ No valid dataframes found to merge from session files.")

def detect_header_and_normalize(file_path):
    # Add to session tracking
    with session_lock:
        session_files.add(file_path)
    
    name_key = get_name_from_filename(os.path.basename(file_path))
    if name_key is None:
        log(f"❌ Skipping file: No valid mapping found for '{file_path}'")
        return

    # Use enhanced header detection
    df = detect_header_row_from_db(file_path, name_key)
    
    # Fallback to old method if new method fails
    if df.empty:
        log(f"⚠️ DB-based header detection failed, falling back to old method")
        df = read_excel_safely(file_path)
    
    if df.empty:
        log(f"❌ Could not read any data from {file_path}")
        return

    mappings = load_mappings_from_db(name_key)
    df.columns = [c.strip() for c in df.columns]
    log(f"⚙️ Normalizing file: {file_path}")
    log(f"🧾 Columns before rename: {list(df.columns)}")
    df.rename(columns=lambda col: mappings.get(col.lower(), col), inplace=True)
    log(f"🧾 Columns after rename: {list(df.columns)}")
    df = split_measurement_columns(df, name_key)

    required_headers = ['color']
    required_lower = {h.lower() for h in required_headers}
    df = df[[col for col in df.columns if col.lower() in required_lower]]

    output_path = os.path.splitext(file_path)[0] + '_normalized.xlsx'
    df.to_excel(output_path, index=False)
    log(f"✅ Normalized file saved to: {output_path}")

    processed_files.add(file_path)
    preprocessing_queue.put(output_path)
    log(f"📥 Added to preprocessing queue: {output_path}")

def process_queue():
    while True:
        file_path = file_queue.get()
        if file_path in processed_files:
            log(f"⚠️ Skipping already processed file: {file_path}")
        else:
            log(f"🌀 Starting normalization for: {file_path}")
            try:
                detect_header_and_normalize(file_path)
            except Exception as e:
                log(f"❌ Error processing file: {file_path} — {e}")
        file_queue.task_done()

def process_queue_and_filter():
    while True:
        if preprocessing_queue.empty():
            time.sleep(1)
            continue
        file_path = preprocessing_queue.get()
        log(f"🧼 Cleaning started for: {file_path}")
        try:
            df = pd.read_excel(file_path)
            cleaned_df = apply_combined_filters(df)
            cleaned_df.to_excel(file_path, index=False)
            log(f"✅ Cleaned (overwritten): {file_path}")
        except Exception as e:
            log(f"❌ Error cleaning {file_path}: {e}")
        preprocessing_queue.task_done()

        # Check if all queues are empty and merge session files
        if file_queue.empty() and preprocessing_queue.empty():
            log("🔍 Queues empty, checking for session merge")
            merge_session_files()

class ExcelHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory and event.src_path.endswith('.xlsx') and '_normalized' not in event.src_path:
            log(f"👀 Watcher saw file created: {event.src_path}")
            file_queue.put(event.src_path)

def start_watcher():
    global observer
    if not os.path.exists(WATCH_FOLDER):
        os.makedirs(WATCH_FOLDER)

    observer = Observer()
    event_handler = ExcelHandler()
    observer.schedule(event_handler, path=WATCH_FOLDER, recursive=False)
    observer.start()

    threading.Thread(target=process_queue, daemon=True).start()
    threading.Thread(target=process_queue_and_filter, daemon=True).start()
    log(f"🚀 Watcher booted and monitoring folder: {WATCH_FOLDER}")

def stop_watcher():
    global observer
    if observer:
        observer.stop()
        observer.join()
        observer = None
        log("🛑 Watcher stopped.")

def clear_queue(q):
    with q.mutex:
        q.queue.clear()
        q.all_tasks_done.notify_all()
        q.unfinished_tasks = 0

def clear_session():
    """Clear current session files"""
    with session_lock:
        session_files.clear()
    log("🗑️ Session cleared.")

def get_session_files():
    """Get current session files"""
    with session_lock:
        return session_files.copy()